import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from rand_delay import RandDelay
from database import Database
import os
from urllib.request import urlretrieve
import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

class CardDescription(RandDelay):


    """
    Класс, в котором реализован парсинг данных из Ссылок, полученный в классе
    Parser, выделяется информация и записывается в Кэш. Затем, после того как
    данные записались в Кэш, данные сохраняются в Exel файл.
    
    """

    def __init__(self, parsed_links, search_query,city):
        self.parsed_links = parsed_links
        self.search_query = search_query
        self.city = city
        self.db_manager = Database()
        self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
        self.process_links()
        

    def process_links(self):
        for url in self.parsed_links:
            self.driver.get(url)
            try:
                # Поиск данных объявления по различным тегам, при помощи get_text, get_price, get_status
                l_id =self.get_text(By.CSS_SELECTOR, '[data-marker="item-view/item-id"]')
                image_link = self.save_images(l_id)
                listing = {
                    'link': url,
                    'search_query': self.search_query,
                    'price': self.get_price(),
                    'description': self.get_text(By.CSS_SELECTOR, '[data-marker="item-view/item-description"]'),
                    'name': self.get_text(By.CLASS_NAME, "style-titleWrapper-Hmr_5"),
                    'city': self.city,
                    'address': self.get_text(By.CLASS_NAME, "style-item-address__string-wt61A"),
                    'id': l_id,
                    'views': self.get_text(By.CSS_SELECTOR, '[data-marker="item-view/total-views"]'),
                    'date': self.get_text(By.CSS_SELECTOR, '[data-marker="item-view/item-date"]'),
                    'images': image_link,
                    'status': self.get_status()
                }
                # Добавление данных в Кэш
                self.db_manager.DML_commands(listing)

            except Exception as e:
                print(f"Ошибка при парсинге {url}: {e}")

            self.slepper()

        #Закрытие браузера, с помощью которого парсятся данные  
        self.driver.quit()
        self.save_to_excel()

    def get_text(self, by_method, element_query):
        try:
            return self.driver.find_element(by_method, element_query).text
        except:
            return "Н/Д"
        
    def get_price(self):
        try:
            return self.driver.find_element(By.CSS_SELECTOR, 'span[data-marker="item-view/item-price"]').get_attribute("content")
        except:
            return "Бесплатно"

    def get_status(self):
        try:
            # Если есть тег closed-warning-block-_5cSD, это значит, что объявление снято с публикации и выставляется статус закрыто
            status = self.driver.find_element(By.CLASS_NAME, "closed-warning-block-_5cSD").text
            return "Закрыто"
        except:
            return "Активно"
    
    #Метод для сохранения изображений в директорию программы /images
    def save_images(self, listing_id):
        try:
            # Извлечение изображений по тегу images-preview-previewImageWrapper-RfThd img
            image_elements = self.driver.find_elements(By.CSS_SELECTOR, '.images-preview-previewImageWrapper-RfThd img')
            
            image_urls = []
            for img in image_elements:
                src = img.get_attribute('src')
                srcset = img.get_attribute('srcset')

                if srcset:
                    candidates = srcset.split(',')

                    for candidate in candidates:
                        url = candidate.split()[0].strip()
                        image_urls.append(url)

                elif src:
                    image_urls.append(src)
            
            image_urls = list(set(image_urls))

            image_folder = f"images/{listing_id}"
            os.makedirs(image_folder, exist_ok=True)

            for idx, url in enumerate(image_urls):
                if url:
                    image_path = f"{image_folder}/{listing_id}_{idx + 1}.jpg"
                    urlretrieve(url, image_path)

            return image_folder 
        except Exception as e:
            print(f"Error occurred: {e}") 
            return "Н/Д"

    #метод для сохранения в Exel файл
    def save_to_excel(self):
        try:
            #Выборка записей из таблицы БД
            connection = sqlite3.connect('cache.db')
            query = "SELECT * FROM listings"
            listings = pd.read_sql_query(query, connection)
            connection.close()

            output_file = 'res_parsing.xlsx'
            df = pd.DataFrame(listings)
            df.to_excel(output_file, index=False)

            wb = load_workbook(output_file)
            ws = wb.active
            
            photo_column = 'images'  

            #добавление ссылок на фотографии в ячейку начиная с H2 до H(n)
            for index, row in df.iterrows():
                photo_path = row[photo_column] 
                cell = f'H{index + 2}'
                ws[cell].hyperlink = photo_path  
                ws[cell].value = "Смотреть фото"  
                ws[cell].style = "Hyperlink"  
                ws[cell].font = Font(underline='single', color='0000FF') 

            wb.save(output_file)

        except Exception as e:
            print(f"Произошла ошибка: {e}")
